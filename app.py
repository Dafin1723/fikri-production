# app.py - Web App Unit Produksi Fikri Production (Flask + SQLite)
# ================================================================
# Pembagian tugas THE FOOL (Minggu ini fokus database & fungsi dasar)
# - Dafin: Struktur awal, config, route home/produk/pesan, upload, download, session
# - Amru: Model database (class Pesanan), inisialisasi db.create_all(), tes dummy data
# - Fikri: Route admin/login/logout, dashboard, update status, flash message & validasi

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

app = Flask(__name__)
app.secret_key = 'rahasia-super-panjang-ubah-ini-pake-os-urandom-bro'  # GANTI NANTI PAKE os.urandom(24)

# CONFIG DATABASE & UPLOAD (Dafin kerjain)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'  # file langsung di root (mudah tes)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}
db = SQLAlchemy(app)

# MODEL DATABASE (Amru kerjain - ini bagian utama database)
class Pesanan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False)
    kontak = db.Column(db.String(100))
    jenis_print = db.Column(db.String(50))
    ukuran = db.Column(db.String(20))
    jumlah = db.Column(db.Integer)
    file_path = db.Column(db.String(255))
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Buat tabel kalau belum ada (Amru kerjain - jalankan sekali aja)
with app.app_context():
    db.create_all()

# Fungsi bantu cek extension file (Dafin kerjain)
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Halaman Home → langsung redirect ke Produk (Dafin kerjain)
@app.route('/')
def home():
    return redirect(url_for('produk'))

# Halaman Katalog Produk (Amru kerjain - template produk.html)
@app.route('/produk')
def produk():
    return render_template('produk.html')

# Halaman Pesan Print + Upload (Fikri kerjain - form & simpan ke DB)
@app.route('/pesan', methods=['GET', 'POST'])
def pesan():
    if request.method == 'POST':
        nama = request.form.get('nama')
        kontak = request.form.get('kontak')
        jenis = request.form.get('jenis_print')
        ukuran = request.form.get('ukuran')
        jumlah = request.form.get('jumlah', type=int)
        file = request.files.get('file')

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            pesanan = Pesanan(nama=nama, kontak=kontak, jenis_print=jenis,
                              ukuran=ukuran, jumlah=jumlah, file_path=file_path)
            db.session.add(pesanan)
            db.session.commit()
            flash('Pesanan berhasil dikirim! Tunggu admin cek ya.', 'success')
        else:
            flash('File tidak valid atau kosong.', 'danger')
        return redirect(url_for('pesan'))
    return render_template('user/index.html')  # atau pesan.html kalau ganti nama

# Halaman Login Admin (Fikri kerjain - hardcode dulu)
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'unitproduksi123':
            session['admin_logged_in'] = True
            session['admin_user'] = username
            flash('Login berhasil! Selamat datang Admin.', 'success')
            return redirect(url_for('admin'))
        else:
            flash('Username atau password salah.', 'danger')
    return render_template('admin/login.html')

# Logout Admin (Fikri kerjain)
@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_user', None)
    flash('Anda telah logout.', 'info')
    return redirect(url_for('admin_login'))

# Proteksi semua route /admin (Dafin kerjain)
@app.before_request
def require_admin_login():
    if request.path.startswith('/admin') and request.path != '/admin/login':
        if not session.get('admin_logged_in'):
            flash('Silakan login terlebih dahulu sebagai admin.', 'warning')
            return redirect(url_for('admin_login'))

# Dashboard Admin (Fikri kerjain - tampil daftar pesanan)
@app.route('/admin')
def admin():
    # Filter berdasarkan status
    status_filter = request.args.get('status', 'all')
    search_query = request.args.get('search', '')
    
    query = Pesanan.query
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search_query:
        query = query.filter(
            (Pesanan.nama.ilike(f'%{search_query}%')) |
            (Pesanan.kontak.ilike(f'%{search_query}%'))
        )
    
    pesanan_list = query.order_by(Pesanan.created_at.desc()).all()
    return render_template('admin/dashboard.html', pesanan=pesanan_list, 
                         status_filter=status_filter, search_query=search_query)

# Update status pesanan (Fikri kerjain)
@app.route('/update/<int:id>', methods=['POST'])
def update(id):
    pesanan = Pesanan.query.get_or_404(id)
    pesanan.status = request.form['status']
    db.session.commit()
    flash('Status berhasil diupdate!', 'info')
    return redirect(url_for('admin'))

# Download file (Dafin kerjain - hanya admin yang bisa akses)
@app.route('/download/<filename>')
def download(filename):
    if not session.get('admin_logged_in'):
        flash('Akses ditolak. Login sebagai admin.', 'danger')
        return redirect(url_for('admin_login'))
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

# Cek Status Pesanan (User bisa cek pake ID)
@app.route('/cek-status', methods=['GET', 'POST'])
def cek_status():
    pesanan = None
    if request.method == 'POST':
        order_id = request.form.get('order_id', type=int)
        pesanan = Pesanan.query.get(order_id)
        if not pesanan:
            flash('Pesanan tidak ditemukan. Cek kembali nomor pesanan Anda.', 'warning')
    return render_template('user/cek_status.html', pesanan=pesanan)

# Export ke Excel
@app.route('/admin/export/excel')
def export_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    pesanan_list = Pesanan.query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pesanan"
    
    # Header styling
    header_fill = PatternFill(start_color="0047FF", end_color="0047FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Nama', 'Kontak', 'Jenis Print', 'Ukuran', 'Jumlah', 'Status', 'Tanggal']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, p in enumerate(pesanan_list, 2):
        ws.cell(row=row_num, column=1, value=p.id)
        ws.cell(row=row_num, column=2, value=p.nama)
        ws.cell(row=row_num, column=3, value=p.kontak)
        ws.cell(row=row_num, column=4, value=p.jenis_print)
        ws.cell(row=row_num, column=5, value=p.ukuran)
        ws.cell(row=row_num, column=6, value=p.jumlah)
        ws.cell(row=row_num, column=7, value=p.status)
        ws.cell(row=row_num, column=8, value=p.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=pesanan_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response

# Export ke PDF
@app.route('/admin/export/pdf')
def export_pdf():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    pesanan_list = Pesanan.query.all()
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    y = 750
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, "Laporan Pesanan Fikri Production")
    y -= 30
    
    p.setFont("Helvetica", 10)
    for pesanan in pesanan_list:
        if y < 50:
            p.showPage()
            y = 750
        p.drawString(50, y, f"ID: {pesanan.id} | {pesanan.nama} | {pesanan.status}")
        y -= 20
    
    p.save()
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=pesanan_{datetime.now().strftime("%Y%m%d")}.pdf'
    return response

# API Statistik untuk Dashboard
@app.route('/admin/api/stats')
def admin_stats():
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    total_pesanan = Pesanan.query.count()
    pending = Pesanan.query.filter_by(status='pending').count()
    proses = Pesanan.query.filter_by(status='proses').count()
    selesai = Pesanan.query.filter_by(status='selesai').count()
    
    return jsonify({
        'total': total_pesanan,
        'pending': pending,
        'proses': proses,
        'selesai': selesai
    })

# Jalankan server (Dafin kerjain - buat folder uploads otomatis)
if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    # Production: gunakan gunicorn, Development: flask run
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
