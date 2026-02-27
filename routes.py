"""Fikri Production – All Routes"""
import os, json
from datetime import datetime
from flask import (render_template, request, redirect, url_for,
    session, flash, jsonify, send_from_directory, make_response, abort)
from werkzeug.utils import secure_filename
from models import (
    get_all_pesanan, get_pesanan_by_id, get_pesanan_by_nomor,
    create_pesanan, update_pesanan_status, delete_pesanan_db,
    generate_nomor_antrian, get_statistics,
    get_all_posters, get_poster_by_id, create_poster, delete_poster_db
)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}
MAX_FILE_SIZE = 20 * 1024 * 1024
MAX_FILES = 10
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'unitproduksi123'


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


def register_routes(app):

    @app.route('/')
    def landing_page():
        posters = get_all_posters()
        return render_template('landing-page.html', posters=posters)

    @app.route('/pesan', methods=['GET', 'POST'])
    def pesan():
        if request.method == 'GET':
            return render_template('form-pemesanan.html')
        try:
            nama = request.form.get('nama', '').strip()
            kontak = request.form.get('kontak', '').strip()
            email = request.form.get('email', '').strip()
            jenis_print = request.form.get('jenis_print', '').strip()
            warna = request.form.get('warna', '').strip()
            ukuran = request.form.get('ukuran', '').strip()
            jenis_kertas = request.form.get('jenis_kertas', '').strip()
            jumlah_str = request.form.get('jumlah', '0').strip()
            tanggal_ambil = request.form.get('tanggal_ambil', '').strip()
            catatan = request.form.get('catatan', '').strip()

            errors = []
            if not nama: errors.append('Nama wajib diisi')
            if not kontak: errors.append('Kontak wajib diisi')
            if not email or '@' not in email: errors.append('Email valid wajib diisi')
            if not jenis_print: errors.append('Jenis print wajib dipilih')
            if not warna: errors.append('Warna wajib dipilih')
            if not ukuran: errors.append('Ukuran wajib dipilih')
            if not jenis_kertas: errors.append('Jenis kertas wajib dipilih')
            if not tanggal_ambil: errors.append('Tanggal ambil wajib diisi')
            try:
                jumlah = int(jumlah_str)
                if jumlah < 1: errors.append('Jumlah minimal 1')
            except:
                errors.append('Jumlah harus berupa angka'); jumlah = 0

            if errors:
                return jsonify({'success': False, 'errors': errors}), 400

            files = request.files.getlist('files[]')
            valid_files = [f for f in files if f and f.filename]
            if len(valid_files) > MAX_FILES:
                return jsonify({'success': False, 'errors': [f'Maksimal {MAX_FILES} file']}), 400

            saved_files = []
            for file in valid_files:
                if not allowed_file(file.filename):
                    return jsonify({'success': False, 'errors': [f'Format tidak diizinkan: {file.filename}']}), 400
                file.seek(0, 2); size = file.tell(); file.seek(0)
                if size > MAX_FILE_SIZE:
                    return jsonify({'success': False, 'errors': [f'{file.filename} melebihi 20MB']}), 400
                orig = secure_filename(file.filename)
                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                fname = f"{ts}_{orig}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
                saved_files.append(fname)

            nomor_antrian = generate_nomor_antrian()
            pesanan_id = create_pesanan({
                'nama': nama, 'kontak': kontak, 'email': email,
                'jenis_print': jenis_print, 'warna': warna, 'ukuran': ukuran,
                'jenis_kertas': jenis_kertas, 'jumlah': jumlah,
                'tanggal_ambil': tanggal_ambil,
                'catatan': catatan if catatan else None,
                'file_paths': json.dumps(saved_files) if saved_files else None,
                'nomor_antrian': nomor_antrian
            })
            return jsonify({'success': True, 'pesanan_id': pesanan_id, 'nomor_antrian': nomor_antrian})

        except Exception as e:
            app.logger.error(f"Order error: {e}")
            return jsonify({'success': False, 'errors': ['Terjadi kesalahan server']}), 500

    @app.route('/receipt/<int:pesanan_id>')
    def receipt(pesanan_id):
        pesanan = get_pesanan_by_id(pesanan_id)
        if not pesanan:
            abort(404)
        return render_template('receipt.html', pesanan=pesanan)

    @app.route('/cek-status', methods=['GET', 'POST'])
    def cek_status():
        pesanan = None; error = None
        if request.method == 'POST':
            nomor = request.form.get('nomor_antrian', '').strip().upper()
            if nomor:
                pesanan = get_pesanan_by_nomor(nomor)
                if not pesanan:
                    error = f'Nomor antrian "{nomor}" tidak ditemukan.'
            else:
                error = 'Masukkan nomor antrian.'
        return render_template('cek-status.html', pesanan=pesanan, error=error)

    @app.route('/uploads/<filename>')
    def uploaded_file(filename):
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

    # ── ADMIN ──

    @app.route('/admin/login', methods=['GET', 'POST'])
    def admin_login():
        if session.get('admin_logged_in'):
            return redirect(url_for('admin_dashboard'))
        if request.method == 'POST':
            u = request.form.get('username', '')
            p = request.form.get('password', '')
            if u == ADMIN_USERNAME and p == ADMIN_PASSWORD:
                session['admin_logged_in'] = True
                flash('Login berhasil!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Username atau password salah!', 'danger')
        return render_template('admin-login.html')

    @app.route('/admin/logout')
    def admin_logout():
        session.clear()
        flash('Anda telah logout.', 'info')
        return redirect(url_for('admin_login'))

    @app.route('/admin/dashboard')
    @admin_required
    def admin_dashboard():
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '').strip()
        pesanan_list = get_all_pesanan(status_filter, search)
        stats = get_statistics()
        return render_template('admin-dashboard.html',
            pesanan_list=pesanan_list, stats=stats,
            status_filter=status_filter, search=search)

    @app.route('/admin/update-status/<int:pesanan_id>', methods=['POST'])
    @admin_required
    def update_status(pesanan_id):
        p = get_pesanan_by_id(pesanan_id)
        if not p: abort(404)
        new_status = request.form.get('status', '')
        if new_status in ['pending', 'proses', 'selesai']:
            update_pesanan_status(pesanan_id, new_status)
            flash(f'Status pesanan {p["nomor_antrian"]} berhasil diperbarui.', 'success')
        else:
            flash('Status tidak valid.', 'danger')
        return redirect(url_for('admin_dashboard',
            status=request.args.get('status', ''), search=request.args.get('search', '')))

    @app.route('/admin/delete/<int:pesanan_id>', methods=['POST'])
    @admin_required
    def delete_pesanan(pesanan_id):
        p = get_pesanan_by_id(pesanan_id)
        if not p: abort(404)
        for fname in p['files']:
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            if os.path.exists(fp): os.remove(fp)
        nomor = p['nomor_antrian']
        delete_pesanan_db(pesanan_id)
        flash(f'Pesanan {nomor} berhasil dihapus.', 'success')
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/export/excel')
    @admin_required
    def export_excel():
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            wb = Workbook(); ws = wb.active; ws.title = 'Pesanan'
            hfill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
            hfont = Font(color='FFFFFF', bold=True, size=11)
            halign = Alignment(horizontal='center', vertical='center')
            border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            headers = ['No','Nomor Antrian','Nama','Email','Kontak','Jenis Print','Warna','Ukuran','Jenis Kertas','Jumlah','Tanggal Ambil','Jml File','Status','Catatan','Dibuat']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hfill; cell.font = hfont; cell.alignment = halign; cell.border = border
            ws.row_dimensions[1].height = 30
            scols = {'pending':'FEF3C7','proses':'DBEAFE','selesai':'D1FAE5'}
            for rn, p in enumerate(get_all_pesanan(), 2):
                row_data = [rn-1, p['nomor_antrian'], p['nama'], p['email'], p['kontak'],
                    p['jenis_print'], p['warna'], p['ukuran'], p['jenis_kertas'], p['jumlah'],
                    p['tanggal_ambil'], p['file_count'], p['status_indo'], p.get('catatan',''), p.get('created_at','')]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=rn, column=col, value=val)
                    cell.border = border; cell.alignment = Alignment(vertical='center')
                    if col == 13:
                        c2 = scols.get(p['status'],'FFFFFF')
                        cell.fill = PatternFill(start_color=c2, end_color=c2, fill_type='solid')
            widths = [5,18,20,25,15,20,15,10,15,10,15,12,12,25,18]
            for i,w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
            ws.insert_rows(1)
            tc = ws.cell(row=1, column=1, value='LAPORAN PESANAN - FIKRI PRODUCTION')
            tc.font = Font(bold=True, size=14, color='4F46E5')
            tc.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            ws.row_dimensions[1].height = 35
            out = io.BytesIO(); wb.save(out); out.seek(0)
            fn = f"pesanan_fikri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            resp = make_response(out.read())
            resp.headers['Content-Disposition'] = f'attachment; filename={fn}'
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return resp
        except Exception as e:
            flash(f'Error export Excel: {e}', 'danger')
            return redirect(url_for('admin_dashboard'))

    @app.route('/admin/export/pdf')
    @admin_required
    def export_pdf():
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            import io
            out = io.BytesIO()
            doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            tstyle = ParagraphStyle('T', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#4F46E5'), alignment=TA_CENTER)
            elements = []
            elements.append(Paragraph('LAPORAN PESANAN - FIKRI PRODUCTION', tstyle))
            elements.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
            headers = ['No','Antrian','Nama','Kontak','Jenis','Warna','Kertas','Jml','Tgl Ambil','Files','Status']
            data = [headers]
            for i, p in enumerate(get_all_pesanan(), 1):
                data.append([str(i), p['nomor_antrian'], p['nama'][:20], p['kontak'],
                    p['jenis_print'][:15], p['warna'], p['jenis_kertas'][:12],
                    str(p['jumlah']), p['tanggal_ambil'], str(p['file_count']), p['status_indo']])
            cw = [1*cm,3*cm,4*cm,3*cm,3.5*cm,2.5*cm,3*cm,1.5*cm,2.5*cm,1.5*cm,2.5*cm]
            t = Table(data, colWidths=cw, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6366F1')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,0),9),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
                ('FONTSIZE',(0,1),(-1,-1),8),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F5F3FF')]),
                ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#E5E7EB')),
                ('TOPPADDING',(0,0),(-1,-1),4),
                ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            elements.append(t)
            doc.build(elements)
            out.seek(0)
            fn = f"pesanan_fikri_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            resp = make_response(out.read())
            resp.headers['Content-Disposition'] = f'attachment; filename={fn}'
            resp.headers['Content-Type'] = 'application/pdf'
            return resp
        except Exception as e:
            flash(f'Error export PDF: {e}', 'danger')
            return redirect(url_for('admin_dashboard'))

    @app.route('/admin/posters')
    @admin_required
    def admin_posters():
        return render_template('admin-posters.html', posters=get_all_posters())

    @app.route('/admin/posters/upload', methods=['POST'])
    @admin_required
    def upload_poster():
        product_name = request.form.get('product_name','').strip()
        title = request.form.get('title','').strip()
        description = request.form.get('description','').strip()
        if not product_name:
            flash('Nama produk wajib diisi.','danger'); return redirect(url_for('admin_posters'))
        file = request.files.get('image')
        if not file or not file.filename:
            flash('Pilih file gambar.','danger'); return redirect(url_for('admin_posters'))
        ext = file.filename.rsplit('.',1)[-1].lower()
        if ext not in {'png','jpg','jpeg','gif','webp'}:
            flash('Format gambar tidak diizinkan.','danger'); return redirect(url_for('admin_posters'))
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f"{ts}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['POSTER_FOLDER'], fname))
        create_poster({'product_name': product_name, 'image_path': fname, 'title': title or None, 'description': description or None})
        flash('Poster berhasil diupload!','success')
        return redirect(url_for('admin_posters'))

    @app.route('/admin/posters/delete/<int:poster_id>', methods=['POST'])
    @admin_required
    def delete_poster(poster_id):
        p = get_poster_by_id(poster_id)
        if p:
            fp = os.path.join(app.config['POSTER_FOLDER'], p['image_path'])
            if os.path.exists(fp): os.remove(fp)
            delete_poster_db(poster_id)
        flash('Poster berhasil dihapus.','success')
        return redirect(url_for('admin_posters'))

    @app.errorhandler(404)
    def not_found(e):
        return render_template('landing-page.html'), 404

    @app.errorhandler(413)
    def too_large(e):
        return jsonify({'success': False, 'errors': ['Total ukuran file terlalu besar.']}), 413
